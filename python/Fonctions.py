from python import app
from supabase import create_client, Client
from flask import Flask, render_template, request, redirect, url_for, session
from dotenv import load_dotenv
import os
from datetime import datetime
import pandas as pd
import io
from flask import send_file, redirect, url_for
from datetime import datetime
# Supabase configuration
load_dotenv()
supabase_url = os.getenv("SUPABASE_URL")
supabase_key = os.getenv("SUPABASE_KEY")
supabase: Client = create_client(supabase_url, supabase_key)

def voirdetailsfonction():
    iddemande=request.form.get('id')
    voirdetails=supabase.table('demandes').select('*').eq('id', iddemande).execute()
    demande=voirdetails.data[0]
    print(demande)
    # Convertir la date de création en format lisible
    
    date_str = demande.get('date_creation')
    if date_str:
            dt = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S.%f")  # format Supabase ISO
            demande['date_creation'] = dt.strftime('%d/%m/%Y %H:%M')
    else:
         demande['date_creation'] = 'N/A'
    return render_template("details.html", demande=demande)

def voirdetailsdirectfonction():
    iddemande=request.form.get('id')
    voirdetails=supabase.table('demandes').select('*').eq('id', iddemande).execute()
    demande=voirdetails.data[0]
    print(demande)
    return render_template("detailsdirect.html", demande=demande)

def changer_statut_fonction():
    iddemande=request.form.get('id')
    statut=request.form.get('statut')
    changerstatut=supabase.table('demandes').update({'statut': statut}).eq('id', iddemande).execute()
    voirdetails=supabase.table('demandes').select('*').eq('id', iddemande).execute()
    demande=voirdetails.data[0]
    # Convertir la date de création en format lisible
    date_str = demande.get('date_creation')
    if date_str:
                    dt = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S.%f")  # format Supabase ISO
                    demande['date_creation'] = dt.strftime('%d/%m/%Y %H:%M')
    else:
                    demande['date_creation'] = 'N/A'
    return render_template("details.html", demande=demande)

def consulter_fonction():
   #selection les demande en attente
    demandes=supabase.table('demandes').select('*').eq('statut', 'en_attente').execute()
    demandes_data = demandes.data   
    for demande in demandes_data:
        # Convertir la date de création en format lisible
        date_str = demande.get('date_creation')
        if date_str:
            dt = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S.%f")  # format Supabase ISO
            demande['date_creation'] = dt.strftime('%d/%m/%Y %H:%M')
        else:
            demande['date_creation'] = 'N/A'
    return render_template("consulter.html", lesdemandes=demandes_data,session=session)

def valide_fonction():
    #selection les demande en attente
    demandes=supabase.table('demandes').select('*').eq('statut', 'approuve').eq('transmis', 'false').execute()
    print(demandes.data)
    demandes_data = demandes.data
    for demande in demandes_data:
        # Convertir la date de création en format lisible
        date_str = demande.get('date_creation')
        if date_str:
            dt = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S.%f")  # format Supabase ISO
            demande['date_creation'] = dt.strftime('%d/%m/%Y %H:%M')
        else:
            demande['date_creation'] = 'N/A'
    return render_template("transmettre.html", lesdemandes=demandes_data,session=session)

def transmettre_fonction():
    iddemande=request.form.get('id')
    #recupere les données de la demande
    demande=supabase.table('demandes').select('*').eq('id', iddemande).execute()
    demande_data = demande.data[0]
    #transmettre la demande
    transmettre=supabase.table('transmis').insert({
        'id': demande_data['id'],
        'titre_demande': demande_data['titre_demande'],
        'categorie': demande_data['categorie'],
        'montant_total': demande_data['montant_total'],
        'role_demandeur': demande_data['role_demandeur'],
        'description': demande_data['description'],
        'departement': demande_data['departement'],
        'date_creation': demande_data['date_creation'],
    }).execute()
    supabase.table('demandes').update({'transmis': 'true'}).eq('id', iddemande).execute()
    return valide_fonction()

def demandes_transmises_fonction():
    departement = session.get('departement')
    demandes_transmises = supabase.table('demandes').select('*').eq('departement', departement).eq('transmis', 'true').execute()
    demandes_data = demandes_transmises.data
    for demande in demandes_data:
        date_str = demande.get('date_creation')
        if date_str:
            dt = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S.%f")  # format Supabase ISO
            demande['date_creation'] = dt.strftime('%d/%m/%Y %H:%M')
        else:
            demande['date_creation'] = 'N/A'
    return render_template("transmis.html", lesdemandes=demandes_data, session=session)

def generer_rapport_excel():
    """
    Génère un rapport Excel avec toutes les demandes de la table transmis
    """
    try:
        demandes_transmises = supabase.table('transmis').select('*').execute()
        
        if not demandes_transmises.data:
            print("Aucune demande trouvée dans la table transmis")
            return redirect(url_for('pagedemande_route'))
        
        print(f"Nombre de demandes trouvées: {len(demandes_transmises.data)}")
        
        data_for_excel = []
        
        for demande in demandes_transmises.data:
            print(f"Traitement de la demande ID: {demande.get('id')}")
            
            date_creation_formatted = 'N/A'
            if demande.get('date_creation'):
                try:
                    dt = datetime.strptime(demande['date_creation'], "%Y-%m-%dT%H:%M:%S.%f")
                    date_creation_formatted = dt.strftime('%d/%m/%Y %H:%M')
                except ValueError:
                    try:
                        dt = datetime.strptime(demande['date_creation'], "%Y-%m-%dT%H:%M:%S")
                        date_creation_formatted = dt.strftime('%d/%m/%Y %H:%M')
                    except:
                        date_creation_formatted = str(demande['date_creation'])
            
            date_modification_formatted = 'N/A'
            if demande.get('date_modification'):
                try:
                    dt = datetime.strptime(demande['date_modification'], "%Y-%m-%dT%H:%M:%S.%f")
                    date_modification_formatted = dt.strftime('%d/%m/%Y %H:%M')
                except ValueError:
                    try:
                        dt = datetime.strptime(demande['date_modification'], "%Y-%m-%dT%H:%M:%S")
                        date_modification_formatted = dt.strftime('%d/%m/%Y %H:%M')
                    except:
                        date_modification_formatted = str(demande['date_modification'])
            
            data_for_excel.append({
                'ID': demande.get('id', ''),
                'Département': demande.get('departement', ''),
                'Rôle Demandeur': demande.get('role_demandeur', ''),
                'Titre de la Demande': demande.get('titre_demande', ''),
                'Catégorie': demande.get('categorie', ''),
                'Description': demande.get('description', ''),
                'Montant (FCFA)': demande.get('montant_total', ''),  
                'Date de Création': date_creation_formatted,
                'Date de Modification': date_modification_formatted,
                'Commentaire Chef': demande.get('commentaire_chef', ''),
                'Commentaire Direction': demande.get('commentaire_direction', ''),
                'Chemin Articles': demande.get('chemin_articles', ''),
                'Pièces Jointes': demande.get('chemin_pieces_jointes', '')
            })
        
        print(f"Données préparées pour Excel: {len(data_for_excel)} lignes")
        
        df = pd.DataFrame(data_for_excel)
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Demandes Transmises', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Demandes Transmises']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        date_actuelle = datetime.now().strftime('%Y%m%d_%H%M%S')
        nom_fichier = f'Rapport_Demandes_Transmises_{date_actuelle}.xlsx'
        
        print(f"Fichier Excel généré: {nom_fichier}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nom_fichier
        )
        
    except Exception as e:
        print(f"Erreur lors de la génération du rapport Excel : {str(e)}")
        import traceback
        traceback.print_exc()
        return redirect(url_for('pagedemande_route'))


def generer_rapport_excel_simple():
    """
    Version alternative avec openpyxl seulement
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        demandes_transmises = supabase.table('transmis').select('*').execute()
        
        if not demandes_transmises.data:
            return redirect(url_for('pagedemande_route'))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Demandes Transmises"
        
        headers = [
            'ID', 'Département', 'Rôle Demandeur', 'Titre de la Demande', 
            'Catégorie', 'Description', 'Montant (FCFA)', 'Date de Création', 
            'Date de Modification', 'Commentaire Chef', 'Commentaire Direction', 
            'Chemin Articles', 'Pièces Jointes'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row, demande in enumerate(demandes_transmises.data, 2):
            date_creation_formatted = 'N/A'
            if demande.get('date_creation'):
                try:
                    dt = datetime.strptime(demande['date_creation'], "%Y-%m-%dT%H:%M:%S.%f")
                    date_creation_formatted = dt.strftime('%d/%m/%Y %H:%M')
                except:
                    try:
                        dt = datetime.strptime(demande['date_creation'], "%Y-%m-%dT%H:%M:%S")
                        date_creation_formatted = dt.strftime('%d/%m/%Y %H:%M')
                    except:
                        date_creation_formatted = str(demande['date_creation'])
            
            date_modification_formatted = 'N/A'
            if demande.get('date_modification'):
                try:
                    dt = datetime.strptime(demande['date_modification'], "%Y-%m-%dT%H:%M:%S.%f")
                    date_modification_formatted = dt.strftime('%d/%m/%Y %H:%M')
                except:
                    try:
                        dt = datetime.strptime(demande['date_modification'], "%Y-%m-%dT%H:%M:%S")
                        date_modification_formatted = dt.strftime('%d/%m/%Y %H:%M')
                    except:
                        date_modification_formatted = str(demande['date_modification'])
            
            data_row = [
                demande.get('id', ''),
                demande.get('departement', ''),
                demande.get('role_demandeur', ''),
                demande.get('titre_demande', ''),
                demande.get('categorie', ''),
                demande.get('description', ''),
                demande.get('montant_total', ''),  
                date_creation_formatted,
                date_modification_formatted,
                demande.get('commentaire_chef', ''),
                demande.get('commentaire_direction', ''),
                demande.get('chemin_articles', ''),
                demande.get('chemin_pieces_jointes', '')
            ]
            
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col, value=value)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        date_actuelle = datetime.now().strftime('%Y%m%d_%H%M%S')
        nom_fichier = f'Rapport_Demandes_Transmises_{date_actuelle}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nom_fichier
        )
        
    except Exception as e:
        print(f"Erreur lors de la génération du rapport Excel : {str(e)}")
        import traceback
        traceback.print_exc()
        return redirect(url_for('pagedemande_route'))
        